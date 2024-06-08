import re
import openpyxl

class Field:
    def __init__(self, bitRange, name):
        self.name = name
        self.description = f"{name} field"
        self.bitRange = bitRange
        
class Register:
    def __init__(self, addressOffset, name):
        self.name = name
        self.description = f"{name} register"
        self.addressOffset = addressOffset
        self.access = "RW"
        self.resetValue = "0x0"
        self.fields = []
        self.dim = None
        
    def add_field(self, field):
        self.fields.append(field)
        
class Peripheral:
    def __init__(self, name, baseAddress, size):
        self.name = name
        self.description = f"{name} peripheral"
        self.baseAddress = baseAddress
        self.size = size
        self.registers = []
        
    def add_register(self, register):
        self.registers.append(register)

def make_field_instance(worksheet, row_index, column_index, max_row):
    fields = []
    for row in worksheet.iter_rows(min_row=row_index, max_row=max_row):
        if row[column_index].value is not None and row[column_index].value != '':
            fields.append(Field(row[column_index-1].value, row[column_index].value))
    return fields
                
def make_register_instance(worksheet):
    registers = []
    merged_cells = worksheet.merged_cells.ranges
    for row in worksheet.iter_rows(min_row=4):
        if row[0].value is None or row[0].value == '':
            continue
        register = Register(row[1].value, row[0].value)
        register.resetValue = row[6].value
        max_row = row[0].row
        for merged_cell in merged_cells:
            if merged_cell.min_row == row[0].row:
                max_row = merged_cell.max_row
                register.dim = str(max_row - row[0].row + 1)
                break
        fields = make_field_instance(worksheet, row[0].row+1, 2, max_row)
        for field in fields:
            register.add_field(field)
        registers.append(register)
    return registers
    
def make_peripheral_instance(worksheet):
    peripheral = Peripheral(worksheet.title, worksheet["B1"].value, worksheet["D1"].value)
    registers = make_register_instance(worksheet)
    for register in registers:
        peripheral.add_register(register)
    return peripheral

# 예시: 엑셀 파일에서 Peripheral 인스턴스 생성
workbook = openpyxl.load_workbook("path/to/your/excel_file.xlsx")
peripherals = []
for worksheet in workbook:
    peripheral = make_peripheral_instance(worksheet)
    peripherals.append(peripheral)